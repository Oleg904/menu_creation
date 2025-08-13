import os
import datetime
import shutil

from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter.messagebox import showinfo


home_dir = os.path.expanduser("~")

num_week_day = 6    # начальная строка для определения недели и дня недели

nutrition_calendar = ''     # путь к файлу календаря питания

month_increase = 0   # прибавление для считывания месяца

file_menu = ''      # путь к файлу типового меню

dates_day_menu = {}     # дни меню с соответствующими им датами

how_day_menu = 0    # сколькидневное меню

def main_window():
    def open_file_calendar():   # открытие файла календаря питания
        global nutrition_calendar
        nutrition_calendar = filedialog.askopenfilename(filetypes=(("EXCEL", ".xlsx"),))
        if nutrition_calendar == '':    # если окно выбора файла было просто закрыто
            showinfo(title="Информация", message="Вы не выбрали календарь питания. Выберите его заново.")
            return
        workbook3 = load_workbook(nutrition_calendar, read_only=True)  # выбор календаря
        sheet3 = workbook3.active  # выбор активного листа в календаре
        if type(sheet3.cell(row=1, column=12).value) != str or 'календарь питания' not in sheet3.cell(row=1, column=12).value.lower():     # проверка того, является ли выбранный файл календарём питания
            showinfo(title="Информация", message="Вы выбрали неверный файл, заново выберите календарь питания")
            return
        # how_much_is_the_daily_menu(sheet3)
        workbook3.close()
        if nutrition_calendar != '':    # если файл календаря был выбран
            showinfo(title="Информация", message="Теперь выберите типовое меню.")

    # открытие файла типового меню
    def open_file_menu():
        global file_menu
        file_menu = filedialog.askopenfilename(filetypes=(("EXCEL", ".xlsx"),))
        if file_menu =='':
            showinfo(title="Информация", message="Вы не выбрали файл типового меню, выберите его заново.")
        elif nutrition_calendar == '':
            showinfo(title="Информация", message="Вы не выбрали календарь питания. Сначала выберите его, а затем заново файл типового меню.")
        else:
            menu_processing()

    def finish():  # закрытие окна
        root.destroy()  # ручное закрытие окна и всего приложения

    # создается главное окно
    root = Tk()
    root.title("Создание меню")
    root.geometry("500x450")

    # вывод текста в основном окне
    label = ttk.Label(root, text="Выберете календарь питания, \nа затем файл типового меню", font=("Arial", 16))
    label.place(relx= 0.5, rely= 0.2, anchor=CENTER)

    # создание кнопки выбора типового меню
    btn = ttk.Button(text="Выбрать меню", command=open_file_menu)
    btn.place(relx= 0.5, rely= 0.5,height=40, width=180, anchor=CENTER)

    # создание кнопки выбора календаря питания
    btn2 = ttk.Button(text="Выбрать календарь питания", command=open_file_calendar)
    btn2.place(relx= 0.5, rely= 0.35,height=40, width=180, anchor=CENTER)

    label2 = ttk.Label(root, text="Создание файлов может занять до одной минуты.\nПрограмма сообщит, когда закончит.\nПо завершении, меню создадутся на Рабочем столе в папке 'Менюшки'",foreground="#126b62", font=("Arial", 10))
    label2.place(relx= 0.5, rely= 0.7, anchor=CENTER)

    # подпись внизу главного окна
    label3 = ttk.Label(root, text="By: Макаров Олег Николаевич МКОУ СОШ №11 г.Тавда", font=("Arial", 8))
    label3.place(relx= 0.3, rely= 0.95, anchor=CENTER)

    root.resizable(False, False)  # запрет изменения размеров окна

    root.protocol("WM_DELETE_WINDOW", finish)

    root.iconbitmap('files/image.ico')

    root.mainloop()


def dates_menu(day, month, year):   # составление списка дней меню с соответствующими им датами
    global dates_day_menu
    global month_increase
    dates_day_menu = {}
    workbook3 = load_workbook(nutrition_calendar, read_only=True)
    sheet3 = workbook3.active
    if month < 6:   # если месяц утверждения меню меньше июня
        month_increase = 3
    elif month > 6 and sheet3.cell(row=9,column=1).value == 'июнь':     # если месяц утверждения июнь
        month_increase = 1
    elif month > 6 and sheet3.cell(row=9,column=1).value != 'июнь' and month == 8:  # если месяц утверждения больше июня
        month_increase = 0
        month = 9
        day = 1
    while True:
        if sheet3.cell(row=month + month_increase, column=day + 1).value is not None and sheet3.cell(row=month + month_increase, column=day + 1).value != 0 and day <= 31:  # если день не праздничный и невыходной или не проставлен 0, а также число не более 31
            date_menu = [year, month, day]  # дата из календаря
            dates_day_menu.setdefault(sheet3.cell(row=month + month_increase, column=day + 1).value, [])
            date_of_month = datetime.date(*date_menu)
            dates_day_menu[sheet3.cell(row=month + month_increase, column=day + 1).value].append(date_of_month.strftime("%Y-%m-%d"))
            day += 1
        elif sheet3.cell(row=month + month_increase, column=day + 1).value is None or sheet3.cell(row=month + month_increase, column=day + 1).value == 0 and day <= 31:     # если день праздничный, выходной или если проставлен 0, а также число не более 31
            day += 1
        if day == 32:   # если счетчик дней стал больше 31, то счётчик дней сбрасывается до 1, а месяц прибавляется на 1
            day = 1
            month += 1
        if month == 6 and day == 1 or month == 13 and day == 1:     # если конец учебного года либо конец календарного года, то цикл завершается
            break
    dates_day_menu = dict(sorted(dates_day_menu.items()))   # сортировка списка дней меню и соответствующих им дат
    workbook3.close()


def cycle(row_of_sheet, sheet, sheet2):     # функция вставки ячеек в ежедневные меню
    row_day_menu = 4    # строка начала вставки в ежедневное меню
    while True:
        if str(sheet.cell(row=row_of_sheet, column=4).value) == "Итого" or str(sheet.cell(row=row_of_sheet, column=4).value) == "итого":     # пропуск строки Итого
            sheet2.cell(row=row_day_menu, column=2).value = sheet.cell(row=row_of_sheet, column=4).value
            sheet2.cell(row=row_day_menu, column=5).value = None  # выход грамм
            sheet2.cell(row=row_day_menu, column=6).value = None  # цена
            sheet2.cell(row=row_day_menu, column=7).value = None  # калорийность
            sheet2.cell(row=row_day_menu, column=8).value = None  # белки
            sheet2.cell(row=row_day_menu, column=9).value = None  # жиры
            sheet2.cell(row=row_day_menu, column=10).value = None  # углеводы
            row_of_sheet += 1
            row_day_menu += 1
            continue
        sheet2.cell(row=row_day_menu,column=1).value = sheet.cell(row=row_of_sheet,column=3).value     # прием пищи
        sheet2.cell(row=row_day_menu, column=2).value = sheet.cell(row=row_of_sheet, column=4).value     # раздел
        sheet2.cell(row=row_day_menu, column=3).value = sheet.cell(row=row_of_sheet, column=11).value     # номер рецепта
        sheet2.cell(row=row_day_menu, column=4).value = sheet.cell(row=row_of_sheet, column=5).value     # наименование блюда
        sheet2.cell(row=row_day_menu, column=5).value = sheet.cell(row=row_of_sheet, column=6).value  # выход грамм
        sheet2.cell(row=row_day_menu, column=6).value = sheet.cell(row=row_of_sheet, column=12).value  # цена
        sheet2.cell(row=row_day_menu, column=7).value = sheet.cell(row=row_of_sheet, column=10).value  # калорийность
        sheet2.cell(row=row_day_menu, column=8).value = sheet.cell(row=row_of_sheet, column=7).value  # белки
        sheet2.cell(row=row_day_menu, column=9).value = sheet.cell(row=row_of_sheet, column=8).value  # жиры
        sheet2.cell(row=row_day_menu, column=10).value = sheet.cell(row=row_of_sheet, column=9).value  # углеводы
        if str(sheet.cell(row=row_of_sheet, column=3).value) == "Итого за день:" or str(sheet.cell(row=row_of_sheet, column=3).value) == "итого за день:":
            global num_week_day
            num_week_day = row_of_sheet + 1     # задание начальной строки для нового дня
            sheet2.cell(row=row_day_menu, column=5).value = None  # выход грамм
            sheet2.cell(row=row_day_menu, column=6).value = None  # цена
            sheet2.cell(row=row_day_menu, column=7).value = None  # калорийность
            sheet2.cell(row=row_day_menu, column=8).value = None  # белки
            sheet2.cell(row=row_day_menu, column=9).value = None  # жиры
            sheet2.cell(row=row_day_menu, column=10).value = None  # углеводы
            break
        else:   # спуск на строку ниже в типовом и ежедневном меню
            row_day_menu += 1
            row_of_sheet += 1


def menu_creation_cycle(school_name, current_date, sheet):  # цикл записи ежедневных меню
    global num_week_day
    num_week_day = 6    # сброс начальной строки для определения недели и дня недели
    for key, value in dates_day_menu.items():     # цикл исходя из календаря питания
        current_date = datetime.datetime.strptime(value[0], "%Y-%m-%d").date()      # создание даты для вставки в ежедневное меню и присвоение имени файла меню
        workbook2 = load_workbook("files/shablon.xlsx")  # открытие шаблона
        sheet2 = workbook2.active  # выбор активного листа
        sheet2.cell(row=1, column=2).value = school_name  # вставка наименования учреждения в ежедневное меню
        sheet2.cell(row=1, column=10).value = current_date.strftime("%d.%m.%Y")  # вставка даты в ежедневное меню
        cycle(num_week_day, sheet, sheet2)
        workbook2.save(
            f"{home_dir}/Desktop/Менюшки/{current_date.strftime("%Y-%m-%d")}-sm.xlsx")  # сохранение файла ежедневного меню
        for dates in value:      # копирование текущего дня меню на соответствующие ему даты
            if datetime.datetime.strptime(dates, "%Y-%m-%d").date() == current_date:    # если текущее значение даты является первым в списке, то пропустить его, т.к. оно уже использовалось
                continue
            print(dates)
            shutil.copy(f"{home_dir}/Desktop/Менюшки/{current_date.strftime("%Y-%m-%d")}-sm.xlsx", f"{home_dir}/Desktop/Менюшки/{datetime.datetime.strptime(dates, "%Y-%m-%d").date().strftime("%Y-%m-%d")}-sm.xlsx")



def menu_processing():
    try:  # проверка на наличие
        start_date = []  # дата начала действия типового меню
        if not os.path.exists(f"{home_dir}/Desktop/Менюшки"):   # проверка наличия папки с ежедневными меню и создание в случае отсутствия
            os.mkdir(f"{home_dir}/Desktop/Менюшки")
        if len(os.listdir(f"{home_dir}/Desktop/Менюшки")) == 0:
            workbook = load_workbook(file_menu, read_only=True)     # выбор файла типового меню
            sheet = workbook.active     # выбор активного листа
            if type(sheet.cell(row=2, column=1).value) != str or 'типовое' not in sheet.cell(row=2, column=1).value.lower():  # проверка того, является ли выбранный файл календарём питания
                showinfo(title="Информация", message="Вы выбрали неверный файл, заново выберите типовое меню")
                return
            school_name = sheet.cell(row=1,column=3).value  # наименование учреждения
            # составление даты начала
            start_date.append(sheet.cell(row=3,column=10).value)     # год
            start_date.append(sheet.cell(row=3,column=9).value)     # месяц
            start_date.append(sheet.cell(row=3,column=8).value)     # день
            date = datetime.date(*start_date)
            current_date = date     # текущая дата начала типового меню
            dates_menu(start_date[2], start_date[1], start_date[0])
            menu_creation_cycle(school_name, current_date, sheet)
            workbook.close()
            showinfo(title="Информация", message="Файлы меню созданы. При необходимости, скорректируйте даты на ежедневных меню. Программу можно закрыть.")
        else:
            showinfo(title="Информация", message="В папке содержатся старые файлы меню. Эти файлы будут перезаписаны.")
            shutil.rmtree(f"{home_dir}/Desktop/Менюшки")
            os.mkdir(f"{home_dir}/Desktop/Менюшки")
            menu_processing()
    except BaseException as errors:
        if 'WinError 32' in str(errors):
            showinfo(title="Информация", message="Закройте файл меню и заново выберите файл типового меню.")
        elif ValueError:
            showinfo(title="Информация", message="В типовом меню не заполнена или введена некорректная дата. Пожалуйста, скорректируйте дату.")
            workbook.close()
        else:
            showinfo(title="Информация", message=str(errors))


main_window()